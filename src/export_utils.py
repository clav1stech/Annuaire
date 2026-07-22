"""Excel export helpers."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from time import sleep
import textwrap
from typing import Callable, Mapping
import unicodedata
from urllib.parse import quote as _url_quote

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


@dataclass
class ColSpec:
    """Column descriptor for name-search structured Excel export."""

    key: str             # key in the result row dict
    header: str          # column header displayed in row 2
    category: str        # category label displayed in row 1
    is_pappers: bool = False  # if True render cell as a Pappers hyperlink


def _sanitize_sheet_name(name: str) -> str:
    sanitized = name[:31]
    for bad in [":", "\\", "/", "?", "*", "[", "]"]:
        sanitized = sanitized.replace(bad, "_")
    return sanitized


def _build_grouped_issue_sheet(
    *,
    siret_overview: pd.DataFrame,
    issue_masks: list[tuple[str, pd.Series]],
) -> pd.DataFrame:
    """Build a grouped issue sheet with one leading issue column."""
    issue_column = "probleme_siret"
    overview_columns = list(siret_overview.columns)
    ordered_columns = [issue_column] + overview_columns

    parts: list[pd.DataFrame] = []

    for issue_label, mask in issue_masks:
        selected = siret_overview.loc[mask].copy()
        if selected.empty:
            continue
        selected.insert(0, issue_column, issue_label)
        parts.append(selected.reindex(columns=ordered_columns))

    if not parts:
        return pd.DataFrame(columns=ordered_columns)
    return pd.concat(parts, ignore_index=True)


def _build_anomalies_sheet(
    *,
    siret_overview: pd.DataFrame,
    status_masks: dict[str, pd.Series],
    input_export_columns: list[str],
    missing_input_rows: pd.DataFrame | None = None,
) -> pd.DataFrame:
    """Build anomalies sheet with Motif + input columns only."""
    motif_column = "Motif"
    selected_input_columns: list[str] = []
    seen: set[str] = set()
    for col in input_export_columns:
        if col in seen:
            continue
        seen.add(col)
        selected_input_columns.append(col)

    overview_existing = set(siret_overview.columns)
    if missing_input_rows is not None and not missing_input_rows.empty:
        missing_existing = {
            col for col in missing_input_rows.columns if col != "motif_exclusion_siret"
        }
        selected_input_columns = [col for col in selected_input_columns if col in overview_existing & missing_existing]
    else:
        selected_input_columns = [col for col in selected_input_columns if col in overview_existing]

    ordered_columns = [motif_column] + selected_input_columns
    parts: list[pd.DataFrame] = []

    if missing_input_rows is not None and not missing_input_rows.empty:
        missing_df = missing_input_rows.copy()
        if "motif_exclusion_siret" in missing_df.columns:
            motif_values = (
                missing_df["motif_exclusion_siret"]
                .fillna("")
                .astype(str)
                .str.strip()
                .replace("", "Identifiant vide ou egal a 0")
            )
        else:
            motif_values = pd.Series(
                ["Identifiant vide ou egal a 0"] * len(missing_df),
                index=missing_df.index,
            )
        missing_df = missing_df.reindex(columns=selected_input_columns)
        missing_df.insert(0, motif_column, motif_values)
        parts.append(missing_df)

    for motif_label, mask in [
        ("Identifiant sans correspondance dans SIRENE", status_masks["not_found"]),
        ("Identifiant invalide", status_masks["invalid"]),
    ]:
        selected = siret_overview.loc[mask].copy()
        if selected.empty:
            continue
        selected = selected.reindex(columns=selected_input_columns)
        selected.insert(0, motif_column, motif_label)
        parts.append(selected)

    if not parts:
        return pd.DataFrame(columns=ordered_columns)
    return pd.concat(parts, ignore_index=True).reindex(columns=ordered_columns)


def _build_doublons_sheet(
    *,
    siret_overview: pd.DataFrame,
    input_export_columns: list[str],
) -> pd.DataFrame:
    """Build duplicates sheet with explicit SIRET/SIREN duplicate motifs."""
    motif_column = "Motif"

    selected_columns: list[str] = []
    seen: set[str] = set()
    for col in input_export_columns:
        if col in siret_overview.columns and col not in seen:
            selected_columns.append(col)
            seen.add(col)

    for col in [
        "siret_entree",
        "siret_normalise",
        "identifiant_recherche",
        "siren",
        "siret_doublon_entree",
        "siren_doublon_entree",
    ]:
        if col in siret_overview.columns and col not in seen:
            selected_columns.append(col)
            seen.add(col)

    normalized = (
        siret_overview.get("siret_normalise", pd.Series(index=siret_overview.index, dtype=str))
        .fillna("")
        .astype(str)
        .str.strip()
    )
    lookup_identifier = (
        siret_overview.get("identifiant_recherche", normalized)
        .fillna("")
        .astype(str)
        .str.strip()
    )
    if "siret_format_valide" in siret_overview.columns:
        valid_format_mask = _bool_from_overview_series(siret_overview["siret_format_valide"])
    else:
        valid_format_mask = lookup_identifier.ne("")
    status_norm = (
        siret_overview.get("siret_status", pd.Series(index=siret_overview.index, dtype=str))
        .fillna("")
        .astype(str)
        .map(_normalize_text)
    )
    eligible_duplicates_mask = valid_format_mask & ~status_norm.str.startswith("inval")

    if "siret_doublon_entree" in siret_overview.columns:
        duplicate_siret_mask = (
            _bool_from_overview_series(siret_overview["siret_doublon_entree"])
            & eligible_duplicates_mask
        )
    else:
        siret_keys = lookup_identifier[eligible_duplicates_mask & lookup_identifier.str.len().eq(14)]
        siret_dup_counts = siret_keys[siret_keys.ne("")].value_counts()
        duplicate_siret_mask = lookup_identifier.map(siret_dup_counts).fillna(0).astype("int64").gt(1)

    if "siren_doublon_entree" in siret_overview.columns:
        duplicate_siren_mask = (
            _bool_from_overview_series(siret_overview["siren_doublon_entree"])
            & eligible_duplicates_mask
        )
    else:
        siren_keys = pd.Series("", index=siret_overview.index, dtype=str)
        valid_lookup_mask = eligible_duplicates_mask & lookup_identifier.ne("")
        siren_keys.loc[valid_lookup_mask & lookup_identifier.str.len().eq(14)] = (
            lookup_identifier.loc[valid_lookup_mask & lookup_identifier.str.len().eq(14)].str[:9]
        )
        siren_keys.loc[valid_lookup_mask & lookup_identifier.str.len().eq(9)] = (
            lookup_identifier.loc[valid_lookup_mask & lookup_identifier.str.len().eq(9)]
        )
        siren_dup_counts = siren_keys[siren_keys.ne("")].value_counts()
        duplicate_siren_mask = siren_keys.map(siren_dup_counts).fillna(0).astype("int64").gt(1)

    ordered_columns = [motif_column] + selected_columns

    siret_sort_key = lookup_identifier.where(lookup_identifier.str.len().eq(14), normalized)
    siren_sort_key = pd.Series("", index=siret_overview.index, dtype=str)
    siren_sort_key.loc[lookup_identifier.str.len().eq(14)] = (
        lookup_identifier.loc[lookup_identifier.str.len().eq(14)].str[:9]
    )
    siren_sort_key.loc[lookup_identifier.str.len().eq(9)] = (
        lookup_identifier.loc[lookup_identifier.str.len().eq(9)]
    )
    siren_sort_key = siren_sort_key.where(siren_sort_key.ne(""), normalized)
    parts: list[pd.DataFrame] = []
    for motif_label, mask, motif_sort_key in [
        ("Doublon SIRET", duplicate_siret_mask, siret_sort_key),
        ("Doublon SIREN", duplicate_siren_mask, siren_sort_key),
    ]:
        selected = siret_overview.loc[mask].copy()
        if selected.empty:
            continue
        selected["_tmp_sort_key"] = motif_sort_key.loc[selected.index]
        selected["_tmp_source_idx"] = selected.index
        selected = selected.sort_values(
            by=["_tmp_sort_key", "_tmp_source_idx"],
            ascending=[True, True],
            kind="mergesort",
        )
        selected = selected.drop(columns=["_tmp_sort_key", "_tmp_source_idx"], errors="ignore")
        selected = selected.reindex(columns=selected_columns)
        selected.insert(0, motif_column, motif_label)
        parts.append(selected.reindex(columns=ordered_columns))

    if not parts:
        return pd.DataFrame(columns=ordered_columns)
    return pd.concat(parts, ignore_index=True).reindex(columns=ordered_columns)


def build_export_sheets(
    siret_overview: pd.DataFrame,
    input_export_columns: list[str],
    missing_input_siret_count: int = 0,
    missing_input_rows: pd.DataFrame | None = None,
    country_filter_applied: bool = False,
    france_input_count: int | None = None,
    non_france_input_count: int | None = None,
    unknown_country_input_count: int | None = None,
    non_france_valid_siret_included_count: int | None = None,
    include_non_france_valid_siret_enabled: bool = False,
    total_input_rows: int | None = None,
) -> dict[str, pd.DataFrame]:
    """Build workbook sheets for business cleanup workflow."""
    status = siret_overview.get("siret_status", pd.Series(index=siret_overview.index, dtype=str))
    status_masks = _status_masks_from_series(status)
    replacement = (
        siret_overview.get(
            "siret_remplacement_recommande",
            pd.Series(index=siret_overview.index, dtype=str),
        )
        .fillna("")
        .astype(str)
        .str.strip()
    )

    anomalies = _build_anomalies_sheet(
        siret_overview=siret_overview,
        status_masks=status_masks,
        input_export_columns=input_export_columns,
        missing_input_rows=missing_input_rows,
    )
    doublons = _build_doublons_sheet(
        siret_overview=siret_overview,
        input_export_columns=input_export_columns,
    )
    siret_a_cloturer = _build_grouped_issue_sheet(
        siret_overview=siret_overview,
        issue_masks=[
            ("SIRET ferme sans remplacant", status_masks["closed"] & replacement.eq("")),
            ("SIRET radie", status_masks["radiated"]),
        ],
    )

    sheets = {
        "siret_overview": siret_overview,
        "statistiques": build_statistics_sheet(
            siret_overview,
            missing_input_siret_count=missing_input_siret_count,
            country_filter_applied=country_filter_applied,
            france_input_count=france_input_count,
            non_france_input_count=non_france_input_count,
            unknown_country_input_count=unknown_country_input_count,
            non_france_valid_siret_included_count=non_france_valid_siret_included_count,
            include_non_france_valid_siret_enabled=include_non_france_valid_siret_enabled,
            total_input_rows=total_input_rows,
        ),
        "dictionnaire_colonnes": build_column_dictionary_sheet(siret_overview),
        "anomalies": anomalies,
        "doublons": doublons,
        "siret_a_cloturer": siret_a_cloturer,
    }
    return sheets


def build_column_dictionary_sheet(siret_overview: pd.DataFrame) -> pd.DataFrame:
    """Build a French dictionary sheet for report columns up to analysis_status_note."""
    descriptions = {
        "siret_entree": "Valeur SIRET/SIREN telle que fournie dans le fichier d'entrée.",
        "siret_normalise": (
            "Valeur d'entree normalisee (chiffres uniquement, troncature a 14 si necessaire)."
        ),
        "identifiant_recherche": (
            "Identifiant effectivement utilise pour la recherche SIRENE "
            "(SIRET 14 chiffres ou SIREN 9 chiffres selon la route de validation)."
        ),
        "siret_format_valide": "Indique si le format est valide (SIRET 14 + Luhn ou SIREN 9 + Luhn).",
        "siret_doublon_entree": (
            "Oui si un SIRET valide (14 + Luhn) apparaît plusieurs fois dans les lignes analysées."
        ),
        "siren_doublon_entree": (
            "Oui si un SIREN valide (9 + Luhn) apparaît plusieurs fois dans les lignes analysées."
        ),
        "siret_status": (
            "Statut métier de l'identifiant: Actif, Fermé, Radiée, Invalide, Non trouvé. "
            "Fermé = établissement fermé avec au moins un autre actif sur le même SIREN. "
            "Radiée = aucun établissement actif sur le SIREN."
        ),
        "cleaning_action": "Action recommandée pour le nettoyage de la base tiers.",
        "siret_remplacement_recommande": (
            "SIRET recommande pour remplacement. "
            "Vide si aucun remplacement n'est applicable."
        ),
        "analysis_succession_disponible": (
            "Oui/Non uniquement pour les SIRET fermes: indique si un lien de succession officiel est disponible."
        ),
        "analysis_nb_candidats_remplacement": (
            "Nombre de candidats actifs sur le meme SIREN quand il n'y a pas de succession. "
            "0 si aucun candidat pour un SIRET ferme. Vide si non applicable."
        ),
        "analysis_raison_choix_remplacement": (
            "Raison du choix du candidat quand plusieurs etablissements actifs sont possibles. "
            "Vide sinon."
        ),
        "analysis_synthese_remplacement": (
            "Synthese du resultat remplacement pour les SIRET fermes: "
            "Succession, Autre SIRET meme SIREN, ou Aucun. Vide sinon."
        ),
        "analysis_nd_detecte": (
            "Oui si au moins une valeur contient le marqueur [ND] (ou [ ND ]), "
            "souvent utilisé quand certaines données ne sont pas diffusées."
        ),
        "analysis_priority": "Niveau de priorité de traitement (Basse/Moyenne/Haute).",
        "analysis_status_note": "Commentaire d'analyse synthétique de la situation.",
        "analysis_data_applied": "Précise quelles données sont affichées (entrée ou remplaçant).",
        "siret_retenu": (
            "SIRET retenu après traitement. "
            "Pour un SIRET actif : identifiant tel que confirmé par SIRENE "
            "(ou SIRET du siège si l'entrée était un SIREN). "
            "Pour un SIRET fermé avec remplacement : SIRET du remplaçant recommandé. "
            "Vide si invalide, non trouvé, ou fermé sans remplacement disponible."
        ),
        "siren": "SIREN de l'entreprise (9 chiffres).",
        "nic": "NIC de l'établissement (5 chiffres).",
        "denomination_entreprise": "Nom/raison sociale principale de l'entreprise.",
        "etatAdministratifEtablissement": "État administratif de l'établissement (A/F).",
        "etatAdministratifUniteLegale": "État administratif de l'unité légale.",
        "adresse_reconstituee": "Adresse établissement reconstituée en une ligne.",
        "codePostalEtablissement": "Code postal de l'établissement.",
        "libelleCommuneEtablissement": "Commune de l'établissement.",
        "dateDebut": "Date de début de validité de l'établissement.",
        "dateFin": "Date de fin de validité de l'établissement.",
        "total_etablissements": "Nombre total d'établissements pour le SIREN.",
        "total_etablissements_actifs": "Nombre d'établissements actifs pour le SIREN.",
        "total_etablissements_fermes": "Nombre d'établissements fermés pour le SIREN.",
        "historique_records_count": "Nombre de lignes historiques disponibles.",
        "historique_latest_dateDebut": "Dernière dateDebut retrouvée dans l'historique.",
        "historique_latest_dateFin": "Dernière dateFin retrouvée dans l'historique.",
    }

    columns = list(siret_overview.columns)
    if "analysis_status_note" in columns:
        cutoff = columns.index("analysis_status_note") + 1
        selected_columns = columns[:cutoff]
    else:
        selected_columns = columns

    records: list[dict[str, str]] = []
    for col in selected_columns:
        if col in descriptions:
            categorie = "Analyse/Contrôle" if col.startswith("analysis_") or col in {
                "siret_status",
                "cleaning_action",
                "siret_remplacement_recommande",
            } else "Donnée SIRENE"
            if col in {
                "siret_entree",
                "siret_normalise",
                "identifiant_recherche",
                "siret_format_valide",
                "siret_doublon_entree",
                "siren_doublon_entree",
            }:
                categorie = "Contrôle format"
            records.append(
                {
                    "colonne": col,
                    "categorie": categorie,
                    "description": _wrap_cell_text(descriptions[col], width=100),
                }
            )
        else:
            records.append(
                {
                    "colonne": col,
                    "categorie": "Entrée utilisateur",
                    "description": _wrap_cell_text(
                        "Colonne issue du fichier d'entrée utilisateur.",
                        width=100,
                    ),
                }
            )
    return pd.DataFrame(records, columns=["colonne", "categorie", "description"])


def _normalize_text(value: object) -> str:
    text = str(value or "").strip().lower()
    normalized = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in normalized if not unicodedata.combining(ch))


def _wrap_cell_text(value: object, width: int = 95) -> str:
    """Wrap long text with explicit line breaks for better Excel readability."""
    text = str(value or "").strip()
    if not text:
        return ""
    wrapped_lines: list[str] = []
    for line in text.splitlines() or [text]:
        if len(line) <= width:
            wrapped_lines.append(line)
            continue
        wrapped_lines.extend(textwrap.wrap(line, width=width, break_long_words=False))
    return "\n".join(wrapped_lines)


def _status_masks_from_series(status_series: pd.Series) -> dict[str, pd.Series]:
    """Build robust status masks from potentially accent/case/encoding-variant labels."""
    status_norm = status_series.fillna("").astype(str).map(_normalize_text)
    return {
        "not_found": status_norm.str.contains("non trouv", regex=False),
        "invalid": status_norm.str.startswith("inval"),
        "closed": status_norm.str.startswith("ferm"),
        "radiated": status_norm.str.startswith("radi"),
        "active": status_norm.str.startswith("actif"),
    }


def _bool_from_overview_series(series: pd.Series) -> pd.Series:
    if series.dtype == bool:
        return series.fillna(False)
    normalized = series.fillna("").astype(str).map(_normalize_text)
    return normalized.isin({"oui", "true", "1", "yes"})


def _stat_row(
    section: str,
    indicator: str,
    value: int | None,
    total: int,
    rule: str,
    part_value: float | None = None,
    include_ratio: bool = True,
) -> dict[str, object]:
    if value is None:
        return {
            "section": section,
            "indicateur": indicator,
            "valeur": "",
            "part_du_total": "",
            "regle_metier": rule,
        }
    if not include_ratio:
        ratio: object = ""
    else:
        ratio = part_value if part_value is not None else ((value / total) if total > 0 else 0.0)
    return {
        "section": section,
        "indicateur": indicator,
        "valeur": int(value),
        "part_du_total": ratio,
        "regle_metier": rule,
    }


def build_statistics_sheet(
    siret_overview: pd.DataFrame,
    missing_input_siret_count: int = 0,
    country_filter_applied: bool = False,
    france_input_count: int | None = None,
    non_france_input_count: int | None = None,
    unknown_country_input_count: int | None = None,
    non_france_valid_siret_included_count: int | None = None,
    include_non_france_valid_siret_enabled: bool = False,
    total_input_rows: int | None = None,
) -> pd.DataFrame:
    """Build the statistics sheet with subtotal rows and separators."""
    df = siret_overview.copy()
    analyzed_rows = int(len(df))
    missing_count = max(0, int(missing_input_siret_count))
    if country_filter_applied:
        france_count = max(0, int(france_input_count if france_input_count is not None else analyzed_rows + missing_count))
        foreign_count = max(0, int(non_france_input_count if non_france_input_count is not None else 0))
        unknown_country_count = max(
            0, int(unknown_country_input_count if unknown_country_input_count is not None else 0)
        )
        included_non_fr_valid_count = max(
            0,
            int(non_france_valid_siret_included_count if non_france_valid_siret_included_count is not None else 0),
        )
        global_total_count = max(
            0,
            int(
                total_input_rows
                if total_input_rows is not None
                else france_count + foreign_count + unknown_country_count
            ),
        )
        analysis_base_total = france_count + unknown_country_count + included_non_fr_valid_count
    else:
        global_total_count = analyzed_rows + missing_count
        france_count = global_total_count
        foreign_count = 0
        unknown_country_count = 0
        included_non_fr_valid_count = 0
        analysis_base_total = global_total_count

    status = df.get("siret_status", pd.Series(index=df.index, dtype=str)).fillna("").astype(str)
    status_masks = _status_masks_from_series(status)
    replacement = (
        df.get("siret_remplacement_recommande", pd.Series(index=df.index, dtype=str))
        .fillna("")
        .astype(str)
        .str.strip()
    )
    synthese = (
        df.get("analysis_synthese_remplacement", pd.Series(index=df.index, dtype=str))
        .fillna("")
        .astype(str)
        .map(_normalize_text)
    )
    nd_series = _bool_from_overview_series(
        df.get("analysis_nd_detecte", pd.Series(index=df.index, data=False))
    )

    status_not_found_count = int(status_masks["not_found"].sum())
    status_invalid_count = int(status_masks["invalid"].sum())
    status_closed_count = int(status_masks["closed"].sum())
    status_radiated_count = int(status_masks["radiated"].sum())
    status_active_count = int(status_masks["active"].sum())

    closed_without_replacement_count = int((status_masks["closed"] & replacement.eq("")).sum())
    closed_with_succession_count = int(
        (status_masks["closed"] & synthese.eq("succession")).sum()
    )
    closed_with_active_same_siren_count = int(
        (status_masks["closed"] & synthese.eq("autre siret meme siren")).sum()
    )
    closed_with_other_source_count = int(
        (
            status_masks["closed"]
            & synthese.ne("")
            & ~(synthese.eq("succession") | synthese.eq("autre siret meme siren") | synthese.eq("aucun"))
        ).sum()
    )

    eligible_nd_mask = status_masks["active"] | (status_masks["closed"] & replacement.ne(""))
    eligible_nd_total = int(eligible_nd_mask.sum())
    nd_count = int((nd_series & eligible_nd_mask).sum())
    no_nd_count = int(((~nd_series) & eligible_nd_mask).sum())
    normalized_keys = (
        df.get("siret_normalise", pd.Series(index=df.index, dtype=str)).fillna("").astype(str).str.strip()
    )
    lookup_identifier = (
        df.get("identifiant_recherche", normalized_keys).fillna("").astype(str).str.strip()
    )
    if "siret_format_valide" in df.columns:
        valid_format_mask = _bool_from_overview_series(df["siret_format_valide"])
    else:
        valid_format_mask = lookup_identifier.ne("")
    status_norm = (
        df.get("siret_status", pd.Series(index=df.index, dtype=str))
        .fillna("")
        .astype(str)
        .map(_normalize_text)
    )
    eligible_duplicates_mask = valid_format_mask & ~status_norm.str.startswith("inval")

    if "siret_doublon_entree" in df.columns:
        duplicate_siret_mask = (
            _bool_from_overview_series(df["siret_doublon_entree"])
            & eligible_duplicates_mask
        )
    else:
        siret_keys = lookup_identifier[eligible_duplicates_mask & lookup_identifier.str.len().eq(14)]
        siret_dup_counts = siret_keys.value_counts()
        duplicate_siret_mask = lookup_identifier.map(siret_dup_counts).fillna(0).astype("int64").gt(1)
    duplicate_siret_count = int(duplicate_siret_mask.sum())

    siren_keys_for_duplicates = pd.Series("", index=df.index, dtype=str)
    valid_lookup_mask = eligible_duplicates_mask & lookup_identifier.ne("")
    siren_keys_for_duplicates.loc[valid_lookup_mask & lookup_identifier.str.len().eq(14)] = (
        lookup_identifier.loc[valid_lookup_mask & lookup_identifier.str.len().eq(14)].str[:9]
    )
    siren_keys_for_duplicates.loc[valid_lookup_mask & lookup_identifier.str.len().eq(9)] = (
        lookup_identifier.loc[valid_lookup_mask & lookup_identifier.str.len().eq(9)]
    )

    if "siren_doublon_entree" in df.columns:
        duplicate_siren_row_mask = (
            _bool_from_overview_series(df["siren_doublon_entree"])
            & eligible_duplicates_mask
        )
    else:
        siren_dup_counts = siren_keys_for_duplicates[siren_keys_for_duplicates.ne("")].value_counts()
        duplicate_siren_row_mask = (
            siren_keys_for_duplicates.map(siren_dup_counts).fillna(0).astype("int64").gt(1)
        )
    duplicate_siren_row_count = int(duplicate_siren_row_mask.sum())
    duplicate_siren_unique_count = int(
        siren_keys_for_duplicates.loc[duplicate_siren_row_mask & siren_keys_for_duplicates.ne("")]
        .value_counts()
        .gt(1)
        .sum()
    )

    def empty_row() -> dict[str, object]:
        return {
            "section": "",
            "indicateur": "",
            "valeur": "",
            "part_du_total": "",
            "regle_metier": "",
        }

    rows: list[dict[str, object]] = []
    if country_filter_applied:
        country_rows: list[dict[str, object]] = [
            _stat_row(
                "Global",
                "Fournisseurs France",
                france_count,
                global_total_count,
                "Lignes dont la colonne pays vaut FR, FRA ou France.",
            ),
            _stat_row(
                "Global",
                "Fournisseurs Etranger",
                foreign_count,
                global_total_count,
                "Lignes dont la colonne pays est renseignee et n'est pas FR/FRA/France (avant regle optionnelle d'inclusion).",
            ),
            _stat_row(
                "Global",
                "Fournisseurs pays non précisé",
                unknown_country_count,
                global_total_count,
                "Lignes dont la colonne pays n'est pas renseignee.",
            ),
        ]
        if include_non_france_valid_siret_enabled:
            country_rows.append(
                _stat_row(
                    "",
                    "\tdont Hors France retenus (identifiant valide)",
                    included_non_fr_valid_count,
                    foreign_count,
                    "Lignes hors France conservees car identifiant au format valide (SIRET/SIREN, longueur + Luhn).",
                    include_ratio=False,
                )
            )
        country_rows.extend(
            [
                _stat_row("", "Total", global_total_count, global_total_count, "", include_ratio=False),
                empty_row(),
                _stat_row(
                    "Global",
                    "Lignes analysées",
                    analyzed_rows,
                    analysis_base_total,
                    "Lignes conservees pour controle SIRENE (France + pays non precise + hors France retenus, identifiant non vide et different de 0).",
                ),
            ]
        )
        rows.extend(country_rows)
    else:
        rows.extend(
            [
                _stat_row(
                    "Global",
                    "Lignes entrée fichier",
                    global_total_count,
                    global_total_count,
                    "Nombre total de lignes du fichier d'entrée (y compris SIRET vides/égaux à 0).",
                ),
                empty_row(),
                _stat_row(
                    "Global",
                    "Lignes analysées",
                    analyzed_rows,
                    global_total_count,
                    "Lignes conservées pour contrôle SIRENE (identifiant non vide et différent de 0).",
                ),
            ]
        )

    rows.extend(
        [
            _stat_row(
                "",
                "\tdont lignes SIRET valides en doublon",
                duplicate_siret_count,
                analyzed_rows,
                (
                    "Sous-ensemble des lignes analysées: identifiant valide de type SIRET "
                    "(clé normalisée sur 14 chiffres) présent au moins 2 fois."
                ),
                include_ratio=False,
            ),
            _stat_row(
                "",
                "\tdont lignes SIREN valides en doublon",
                duplicate_siren_row_count,
                analyzed_rows,
                (
                    "Sous-ensemble des lignes analysées: identifiant valide de type SIREN "
                    "(clé normalisée sur 9 chiffres) présent au moins 2 fois."
                ),
                include_ratio=False,
            ),
            _stat_row(
                "",
                "\tdont SIREN valides distincts en doublon",
                duplicate_siren_unique_count,
                max(1, analyzed_rows),
                (
                    "Nombre de SIREN valides distincts (clé normalisée sur 9 chiffres) "
                    "apparaissant sur plusieurs lignes analysées."
                ),
                include_ratio=False,
            ),
            empty_row(),
        ]
    )

    rows.extend(
        [
            _stat_row(
                "Qualité entrée",
                "Identifiants absents dans le fichier d'entrée",
                missing_count,
                analysis_base_total,
                "Lignes dont la valeur identifiant est vide ou égale à 0 dans le fichier source.",
            ),
            _stat_row("", "Total", analysis_base_total, analysis_base_total, "", include_ratio=False),
            empty_row(),
            _stat_row(
                "Statut identifiant",
                "Identifiants sans correspondance dans SIRENE",
                status_not_found_count,
                analyzed_rows,
                "Identifiants au format valide (SIRET/SIREN) mais non retrouvés dans stocketablissement.",
            ),
            _stat_row(
                "Statut identifiant",
                "Identifiants invalides",
                status_invalid_count,
                analyzed_rows,
                "Identifiants rejetés au contrôle de format (SIRET/SIREN).",
            ),
            _stat_row(
                "Statut identifiant",
                "SIRET fermés",
                status_closed_count,
                analyzed_rows,
                "Établissement fermé avec au moins un autre établissement actif sur le même SIREN.",
            ),
            _stat_row(
                "Statut identifiant",
                "SIRET radiés",
                status_radiated_count,
                analyzed_rows,
                "Aucun établissement actif retrouvé sur le SIREN.",
            ),
            _stat_row(
                "Statut identifiant",
                "SIRET actifs",
                status_active_count,
                analyzed_rows,
                "SIRET actifs.",
            ),
            _stat_row("", "Total", analyzed_rows, analyzed_rows, "", include_ratio=False),
            empty_row(),
            _stat_row(
                "Fermés et remplacements",
                "Fermés avec succession officielle",
                closed_with_succession_count,
                status_closed_count,
                "Remplacement basé sur un lien de succession SIRENE.",
            ),
            _stat_row(
                "Fermés et remplacements",
                "Fermés avec actif même SIREN",
                closed_with_active_same_siren_count,
                status_closed_count,
                "Pas de succession, remplacement proposé via un établissement actif du même SIREN.",
            ),
            _stat_row(
                "Fermés et remplacements",
                "Fermés avec source autre/non précisée",
                closed_with_other_source_count,
                status_closed_count,
                "Remplacement trouvé mais source non reconnue par les règles standard.",
            ),
            _stat_row(
                "Fermés et remplacements",
                "Fermés sans remplaçant",
                closed_without_replacement_count,
                status_closed_count,
                "SIRET fermés sans proposition de remplacement.",
            ),
            _stat_row("", "Total", status_closed_count, status_closed_count, "", include_ratio=False),
            empty_row(),
            _stat_row(
                "Disponibilité des données",
                "Lignes avec marqueur [ND]",
                nd_count,
                eligible_nd_total,
                (
                    "Population ND = SIRET actifs + SIRET fermés avec remplaçant. "
                    "Au moins un champ contient [ND] (données partiellement non diffusées)."
                ),
            ),
            _stat_row(
                "Disponibilité des données",
                "Lignes sans marqueur [ND]",
                no_nd_count,
                eligible_nd_total,
                (
                    "Population ND = SIRET actifs + SIRET fermés avec remplaçant. "
                    "Aucun marqueur [ND] détecté."
                ),
            ),
            _stat_row("", "Total", eligible_nd_total, eligible_nd_total, "", include_ratio=False),
        ]
    )

    stats_df = pd.DataFrame(
        rows,
        columns=["section", "indicateur", "valeur", "part_du_total", "regle_metier"],
    )
    ratio_values = pd.to_numeric(stats_df["part_du_total"], errors="coerce")
    stats_df["part_du_total"] = ratio_values.where(ratio_values.notna(), "")
    subset_mask = stats_df["indicateur"].fillna("").astype(str).str.startswith("\tdont ")
    stats_df.loc[subset_mask, "part_du_total"] = ""
    return stats_df


ExcelProgressCallback = Callable[[int, str], None]


def to_excel_bytes(
    sheets: Mapping[str, pd.DataFrame],
    progress_callback: ExcelProgressCallback | None = None,
) -> bytes:
    """Serialize multiple sheets to an Excel binary payload."""
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            safe_name = _sanitize_sheet_name(sheet_name)
            export_df = df if df is not None else pd.DataFrame()
            if safe_name == "siret_overview":
                if progress_callback:
                    progress_callback(88, "\u00c9criture des donn\u00e9es Excel...")
                export_df.to_excel(writer, sheet_name=safe_name, index=False, startrow=1)
                ws = writer.sheets[safe_name]
                _style_siret_overview_sheet(
                    ws=ws,
                    df=export_df,
                    progress_callback=progress_callback,
                )
            elif safe_name == "statistiques":
                export_df.to_excel(writer, sheet_name=safe_name, index=False)
                ws = writer.sheets[safe_name]
                _style_statistics_sheet(ws=ws, df=export_df)
            elif safe_name == "dictionnaire_colonnes":
                export_df.to_excel(writer, sheet_name=safe_name, index=False)
                ws = writer.sheets[safe_name]
                _style_dictionary_sheet(ws=ws, df=export_df)
            elif safe_name in {"anomalies", "doublons", "siret_a_cloturer"}:
                export_df.to_excel(writer, sheet_name=safe_name, index=False)
                ws = writer.sheets[safe_name]
                _style_generic_table_sheet(ws=ws, df=export_df, freeze_cell="A2")
            else:
                export_df.to_excel(writer, sheet_name=safe_name, index=False)
        if progress_callback:
            progress_callback(99, "Finalisation du classeur Excel...")
    buffer.seek(0)
    return buffer.getvalue()


def _fallback_output_path(output_path: Path) -> Path:
    """Build a non-conflicting fallback output path."""
    stem = output_path.stem
    suffix = output_path.suffix if output_path.suffix else ".xlsx"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    candidate = output_path.with_name(f"{stem}_{timestamp}{suffix}")
    idx = 1
    while candidate.exists():
        candidate = output_path.with_name(f"{stem}_{timestamp}_{idx}{suffix}")
        idx += 1
    return candidate


def save_excel_file(
    sheets: Mapping[str, pd.DataFrame],
    output_path: Path,
    payload: bytes | None = None,
) -> Path:
    """Write workbook to disk and return final path.

    If the target file is locked, a fallback file name is used.
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)
    final_payload = payload if payload is not None else to_excel_bytes(sheets)

    # A short retry can absorb transient lock/contention from sync tools.
    for _ in range(2):
        try:
            output_path.write_bytes(final_payload)
            return output_path
        except PermissionError:
            sleep(0.5)

    fallback = _fallback_output_path(output_path)
    fallback.write_bytes(final_payload)
    return fallback


def _style_statistics_sheet(ws, df: pd.DataFrame) -> None:
    """Apply simple styling to statistics sheet."""
    if df.empty:
        return

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    section_fill = PatternFill("solid", fgColor="E2F0D9")
    total_fill = PatternFill("solid", fgColor="EDEDED")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    max_col = len(df.columns)
    max_row = len(df) + 1
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border

    section_col = 1
    indicator_col = list(df.columns).index("indicateur") + 1 if "indicateur" in df.columns else None
    percent_col = list(df.columns).index("part_du_total") + 1 if "part_du_total" in df.columns else None
    for row_idx in range(2, max_row + 1):
        row_values = [str(ws.cell(row=row_idx, column=col_idx).value or "").strip() for col_idx in range(1, max_col + 1)]
        is_blank_row = all(value == "" for value in row_values)
        indicator_raw = (
            str(ws.cell(row=row_idx, column=indicator_col).value or "")
            if indicator_col
            else ""
        )
        indicator_value = (
            indicator_raw.strip()
            if indicator_col
            else ""
        )
        is_total_row = indicator_value == "Total"
        is_subset_row = bool(indicator_raw.startswith("\tdont "))

        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if is_blank_row:
                cell.border = Border()
                cell.fill = PatternFill(fill_type=None)
            else:
                cell.border = thin_border
                cell.alignment = left if col_idx not in {3, percent_col} else center
        if is_blank_row:
            ws.row_dimensions[row_idx].height = 10
            continue

        section_value = str(ws.cell(row=row_idx, column=section_col).value or "").strip()
        if section_value:
            ws.cell(row=row_idx, column=section_col).fill = section_fill
            ws.cell(row=row_idx, column=section_col).font = Font(bold=True, color="1F4E78")
        if is_total_row:
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = total_fill
                cell.font = Font(bold=True, color="1F1F1F")
        elif is_subset_row:
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = Font(italic=True, color="4F4F4F")
        if percent_col is not None:
            percent_cell = ws.cell(row=row_idx, column=percent_col)
            if isinstance(percent_cell.value, (int, float)) and not isinstance(percent_cell.value, bool):
                percent_cell.number_format = "0.00%"

    width_overrides = {
        "section": 28,
        "indicateur": 44,
        "valeur": 12,
        "part_du_total": 14,
        "regle_metier": 82,
    }
    for idx, col in enumerate(df.columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width_overrides.get(col, 24)

    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
    ws.freeze_panes = "A2"


def _style_dictionary_sheet(ws, df: pd.DataFrame) -> None:
    """Apply richer formatting to the dictionary sheet."""
    if df.empty:
        return

    _style_generic_table_sheet(ws=ws, df=df, freeze_cell="A2")

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    section_fill = PatternFill("solid", fgColor="FFF2CC")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    max_col = len(df.columns)
    max_row = len(df) + 1
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border

    category_col = list(df.columns).index("categorie") + 1 if "categorie" in df.columns else None
    for row_idx in range(2, max_row + 1):
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = left
        if category_col:
            cat_cell = ws.cell(row=row_idx, column=category_col)
            cat_cell.fill = section_fill
            cat_cell.font = Font(bold=True, color="7F6000")
            cat_cell.alignment = center

    width_overrides = {
        "colonne": 38,
        "categorie": 26,
        "description": 90,
    }
    for idx, col in enumerate(df.columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width_overrides.get(col, 28)

    # Harmonize vertical alignment and increase row height when long wrapped text is present.
    for row_idx in range(2, max_row + 1):
        row_values = [str(ws.cell(row=row_idx, column=col_idx).value or "") for col_idx in range(1, max_col + 1)]
        max_lines = max(1, *[value.count("\n") + 1 for value in row_values])
        ws.row_dimensions[row_idx].height = max(18, min(120, 18 * max_lines))

    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
    ws.freeze_panes = "A2"


def _style_generic_table_sheet(ws, df: pd.DataFrame, freeze_cell: str = "A2") -> None:
    """Apply a clean generic style for simple export sheets."""
    if df.empty:
        return

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    max_col = len(df.columns)
    max_row = len(df) + 1
    for col_idx in range(1, max_col + 1):
        head = ws.cell(row=1, column=col_idx)
        head.fill = header_fill
        head.font = header_font
        head.alignment = center
        head.border = thin_border

    sample_size = min(len(df), 200)
    for row_idx in range(2, max_row + 1):
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = left

    for idx, col in enumerate(df.columns, start=1):
        preview = [str(col)] + ["" if pd.isna(x) else str(x) for x in df[col].head(sample_size)]
        max_len = max(len(v) for v in preview) if preview else 12
        ws.column_dimensions[get_column_letter(idx)].width = max(12, min(55, max_len + 2))

    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
    ws.freeze_panes = freeze_cell


def _column_group(column_name: str, input_columns: set[str]) -> str:
    """Return visual group label for a column."""
    if column_name in input_columns or column_name == "siret_entree":
        return "Input utilisateur"
    if column_name in {
        "siret_normalise",
        "identifiant_recherche",
        "siret_format_valide",
        "siret_doublon_entree",
        "siren_doublon_entree",
    }:
        return "Contr\u00f4les format"
    if (
        column_name.startswith("analysis_")
        or column_name
        in {
            "siret_status",
            "cleaning_action",
            "siret_remplacement_recommande",
        }
    ):
        return "Analyse situation"
    return "Donn\u00e9es brutes SIRENE"


def _style_siret_overview_sheet(
    ws,
    df: pd.DataFrame,
    progress_callback: ExcelProgressCallback | None = None,
) -> None:
    """Apply business-friendly styling to the siret_overview sheet."""
    if df.empty:
        return

    header_row = 2
    first_data_row = 3
    max_col = len(df.columns)
    max_row = first_data_row + len(df) - 1

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    is_large_report = len(df) >= 7000
    group_styles = {
        "Input utilisateur": PatternFill("solid", fgColor="DCE6F1"),
        "Contr\u00f4les format": PatternFill("solid", fgColor="E2F0D9"),
        "Donn\u00e9es brutes SIRENE": PatternFill("solid", fgColor="FCE4D6"),
        "Analyse situation": PatternFill("solid", fgColor="FFE699"),
    }
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    group_font = Font(color="1F1F1F", bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    input_columns = set()
    for col in df.columns:
        if col in {"siret_entree", "siret_normalise"}:
            break
        input_columns.add(col)

    # Group header row (row 1) with merged ranges.
    groups = [_column_group(col, input_columns=input_columns) for col in df.columns]
    segments: list[tuple[int, int, str]] = []
    segment_start = 1
    for idx in range(2, max_col + 1):
        if groups[idx - 1] != groups[idx - 2]:
            segments.append((segment_start, idx - 1, groups[segment_start - 1]))
            segment_start = idx
    segments.append((segment_start, max_col, groups[segment_start - 1]))

    center_across = Alignment(horizontal="centerContinuous", vertical="center", wrap_text=True)
    for start, end, group_name in segments:
        fill = group_styles.get(group_name, PatternFill("solid", fgColor="D9D9D9"))
        for c in range(start, end + 1):
            cell = ws.cell(row=1, column=c, value=group_name if c == start else None)
            cell.fill = fill
            cell.font = group_font
            cell.alignment = center_across
            cell.border = thin_border

    # Header row style.
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border

    # For very large exports, avoid full-cell styling (too slow).
    if not is_large_report:
        stripe_fill = PatternFill("solid", fgColor="F8FBFF")
        row_count = max(1, len(df))
        emit_every = max(150, int(row_count / 25))
        for row_idx in range(first_data_row, max_row + 1):
            striped = (row_idx - first_data_row) % 2 == 1
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if striped:
                    cell.fill = stripe_fill
                cell.alignment = left
                cell.border = thin_border
            if progress_callback and ((row_idx - first_data_row + 1) % emit_every == 0):
                ratio = (row_idx - first_data_row + 1) / row_count
                percent = min(95, int(90 + ratio * 5))
                progress_callback(percent, "Mise en forme du report Excel...")

    # Status coloring for quick triage.
    header_positions = {str(col): idx + 1 for idx, col in enumerate(df.columns)}
    status_col = header_positions.get("siret_status")
    priority_col = header_positions.get("analysis_priority")
    action_col = header_positions.get("cleaning_action")
    replacement_col = header_positions.get("siret_remplacement_recommande")

    status_colors = {
        "Actif": "C6EFCE",
        "Ferm\u00e9": "F8CBAD",
        "Non trouv\u00e9": "D9E1F2",
        "Invalide": "FCE4D6",
        "Radi\u00e9e": "FFF2CC",
    }
    priority_colors = {"Haute": "F4B084", "Moyenne": "FFD966", "Basse": "C6E0B4"}
    replacement_fill = PatternFill("solid", fgColor="E2F0D9")

    row_count = max(1, len(df))
    emit_every = max(150, int(row_count / 20))
    for row_idx in range(first_data_row, max_row + 1):
        if status_col:
            status_value = str(ws.cell(row=row_idx, column=status_col).value or "")
            fill_color = status_colors.get(status_value)
            if fill_color:
                ws.cell(row=row_idx, column=status_col).fill = PatternFill("solid", fgColor=fill_color)
        if priority_col:
            priority_value = str(ws.cell(row=row_idx, column=priority_col).value or "")
            priority_color = priority_colors.get(priority_value)
            if priority_color:
                ws.cell(row=row_idx, column=priority_col).fill = PatternFill(
                    "solid",
                    fgColor=priority_color,
                )
        if action_col:
            ws.cell(row=row_idx, column=action_col).font = Font(bold=True, color="1F4E78")
        if replacement_col:
            replacement_value = str(ws.cell(row=row_idx, column=replacement_col).value or "").strip()
            if replacement_value:
                ws.cell(row=row_idx, column=replacement_col).fill = replacement_fill

        # Keep key analysis cells readable in large mode without styling every cell.
        if is_large_report:
            if status_col:
                ws.cell(row=row_idx, column=status_col).alignment = center
            if priority_col:
                ws.cell(row=row_idx, column=priority_col).alignment = center
            if replacement_col:
                ws.cell(row=row_idx, column=replacement_col).alignment = center
        if progress_callback and ((row_idx - first_data_row + 1) % emit_every == 0):
            ratio = (row_idx - first_data_row + 1) / row_count
            percent = min(98, int(95 + ratio * 3))
            progress_callback(percent, "Colorisation des analyses...")

    # Column widths.
    width_overrides = {
        "analysis_status_note": 55,
        "adresse_reconstituee": 42,
        "denomination_entreprise": 36,
        "cleaning_action": 34,
    }
    for idx, col in enumerate(df.columns, start=1):
        if col in width_overrides:
            width = width_overrides[col]
        else:
            preview_size = 60 if is_large_report else 200
            series_preview = df[col].astype(str).head(preview_size).tolist()
            max_len = max([len(str(col)), *[len(value) for value in series_preview]])
            width = max(11, min(30, max_len + 2))
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(max_col)}{max_row}"
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = True


def to_name_search_excel_bytes(
    result_rows: list[dict],
    col_specs: list[ColSpec],
    sheet_name: str = "resultats_recherche",
) -> bytes:
    """Build a formatted Excel workbook for name-search results.

    Layout:
      Row 1 — category headers (INPUT / ALICE / Candidat N), merged cells per group.
      Row 2 — column headers (dark blue).
      Row 3+ — data; Pappers columns rendered as clickable hyperlinks.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = _sanitize_sheet_name(sheet_name)

    _CAT_FILL = {
        "INPUT": PatternFill("solid", fgColor="DCE6F1"),
        "ALICE": PatternFill("solid", fgColor="E2F0D9"),
    }
    _CAT_FONT = {
        "INPUT": Font(bold=True, color="1F4E78"),
        "ALICE": Font(bold=True, color="375623"),
    }
    _CAND_FILLS = [
        PatternFill("solid", fgColor="FCE4D6"),
        PatternFill("solid", fgColor="FFF2CC"),
        PatternFill("solid", fgColor="EAF0FB"),
        PatternFill("solid", fgColor="F2E0F2"),
    ]
    _CAND_FONTS = [
        Font(bold=True, color="843C24"),
        Font(bold=True, color="7F6000"),
        Font(bold=True, color="1F4E78"),
        Font(bold=True, color="5B1D5B"),
    ]
    thin = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    pappers_font = Font(color="0563C1", underline="single")
    pappers_base = "https://www.pappers.fr/recherche?q="
    n = len(col_specs)

    # ── Row 2: column headers ────────────────────────────────────────────────
    for ci, spec in enumerate(col_specs, 1):
        cell = ws.cell(row=2, column=ci, value=spec.header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin

    # ── Row 3+: data ────────────────────────────────────────────────────────
    stripe = PatternFill("solid", fgColor="F8FBFF")
    for ri, row_data in enumerate(result_rows, 3):
        odd = (ri - 3) % 2 == 1
        for ci, spec in enumerate(col_specs, 1):
            raw = row_data.get(spec.key, "")
            cell = ws.cell(row=ri, column=ci)
            cell.border = thin
            cell.alignment = left
            if odd:
                cell.fill = stripe
            if spec.is_pappers:
                if raw:
                    cell.hyperlink = pappers_base + _url_quote(str(raw), safe="")
                    cell.value = "Recherche"
                    cell.font = pappers_font
            else:
                cell.value = raw if raw != "" and raw is not None else None

    # ── Row 1: category headers with merged cells ────────────────────────────
    segments: list[tuple[int, int, str]] = []
    if col_specs:
        seg_start = 1
        for ci in range(2, n + 1):
            if col_specs[ci - 1].category != col_specs[ci - 2].category:
                segments.append((seg_start, ci - 1, col_specs[seg_start - 1].category))
                seg_start = ci
        segments.append((seg_start, n, col_specs[seg_start - 1].category))

    cand_idx = 0
    for start, end, category in segments:
        if start < end:
            ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        cell = ws.cell(row=1, column=start, value=category)
        if category in _CAT_FILL:
            cell.fill = _CAT_FILL[category]
            cell.font = _CAT_FONT[category]
        else:
            idx = min(cand_idx, len(_CAND_FILLS) - 1)
            cell.fill = _CAND_FILLS[idx]
            cell.font = _CAND_FONTS[idx]
            cand_idx += 1
        cell.alignment = center
        for c in range(start, end + 1):
            ws.cell(row=1, column=c).border = thin

    # ── Column widths ────────────────────────────────────────────────────────
    for ci, spec in enumerate(col_specs, 1):
        h = spec.header.lower()
        if spec.is_pappers:
            w = 12
        elif "denomination" in h or spec.header == "NOM":
            w = 30
        elif "adresse" in h or spec.header == "ADR":
            w = 38
        elif spec.header in ("Priorité", "ZIP", "PAYS", "SIRET"):
            w = 11
        elif "score" in h:
            w = 10
        else:
            preview = [spec.header] + [
                str(r.get(spec.key, "") or "") for r in result_rows[:60]
            ]
            w = max(10, min(45, max(len(v) for v in preview) + 2))
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Freeze panes + autofilter ────────────────────────────────────────────
    ws.freeze_panes = "A3"
    last_row = 2 + len(result_rows)
    ws.auto_filter.ref = f"A2:{get_column_letter(n)}{last_row}"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()






